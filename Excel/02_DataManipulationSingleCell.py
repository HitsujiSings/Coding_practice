from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

wb = load_workbook('balances.xlsx')
ws = wb.active
ws1 = wb.create_sheet()
ws1.title = "New Title"

ws['A4'] = 4
c = ws['A4']
print("c=",c)
d = ws.cell('A4')
print("d.value=", d.value)
e = ws.cell(row = 4, column = 1)
print("e=",e)


wb.save('balances02.xlsx')
