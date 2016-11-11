from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

wb = load_workbook('balances.xlsx')
ws = wb.active

#cell_range = ws['A1':'C2']
#for cell in cell_range:
#    print(cell)

print(ws.rows) #print out the rows between first and last data 


wb.save('balances03.xlsx')
